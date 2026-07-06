import openpyxl, json, sys

files = {
    "INTERESADOS": "C:/Users/adm/Downloads/G1-Bustamante-Sanchez-Dominio_Interesados.xlsx",
    "RECURSOS":    "C:/Users/adm/Downloads/G1-Recursos_por_Actividad_PMBOK8.xlsx",
    "RIESGOS":     "C:/Users/adm/Downloads/G1-gestion_riesgos.xlsx",
}

for label, path in files.items():
    wb = openpyxl.load_workbook(path)
    out = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                rows.append([str(c) if c is not None else '' for c in row])
        out[sh] = rows
    with open(f'D:/DISCO_D/GESTION_PROYECTOS/riskflow-web/server/scripts/dump_{label}.json', 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False)
    print(f"{label}: {len(out)} sheets", file=sys.stderr)
    for sh, rows in out.items():
        print(f"  Sheet '{sh}': {len(rows)} rows", file=sys.stderr)
